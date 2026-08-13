#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
污染清单生成（只读审计，不改迁移）：扫描老库字符串列，命中注入/攻击特征 → 输出 Excel。

命中正则与迁移脚本 _is_polluted 一致（_POLLUTED_RX）：
    union select / select from / sleep( / benchmark( / md5( / updatexml( /
    extractvalue( / ${jndi: / %bf / 0x[hex]{6,} / limit N#

输出：/Users/rs/Documents/污染清单.xlsx（按表分 sheet + 汇总 sheet）

用法：
    .venv/bin/python scripts/gen_pollution_list.py
"""
import re
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from pymysql.cursors import DictCursor

from migrate_old_db_to_shadow import db, OLD_DB

OUT = "/Users/rs/Documents/污染清单.xlsx"

_POLLUTED_RX = re.compile(
    r"union\s+select|select\s+from|sleep\s*\(|benchmark\s*\(|md5\s*\(|"
    r"updatexml\s*\(|extractvalue\s*\(|\$\{jndi:|%bf|limit\s+\d+\s*#|0x[0-9a-fA-F]{6,}",
    re.IGNORECASE)

# 待扫描表：表名 → 需要扫描的字符串列（取全列，含主键列便于定位）
TABLES = {
    "user_address": ["id", "name", "phone", "province", "city", "district",
                     "address_detail", "community_name", "community_code", "house_no"],
    "order": ["order_id", "account_id", "user_name", "user_phone", "province", "city",
              "district", "address_detail", "community_name", "community_code", "house_num",
              "remark"],
    "order_product": ["order_product_id", "product_name", "product_img", "sku_name",
                      "remark", "store_code", "package_code", "imei", "sn"],
    "sys_user": ["account_id", "phone", "nick_name", "real_auth_name"],
}


def match_pattern(s):
    m = _POLLUTED_RX.search(s)
    return m.group(0) if m else None


def _sql_regex():
    """与 _POLLUTED_RX 同义的 SQL REGEXP（参数化传递，避免字符串字面量转义破坏）"""
    return "|".join([
        r"union[ ]+select", r"select[ ]+from", r"sleep[ ]*\(", r"benchmark[ ]*\(",
        r"md5[ ]*\(", r"updatexml[ ]*\(", r"extractvalue[ ]*\(", r"%bf",
        r"limit[ ]+[0-9]+[ ]*#", r"0x[0-9a-fA-F]{6,}",
    ])


def scan(cur, table, cols):
    """返回 [(id值, 命中字段, 命中特征, 命中列内容样例)]"""
    hits = []
    pk = cols[0]
    rx = _sql_regex()
    for col in cols[1:]:
        cur.execute(
            f"SELECT `{pk}`, `{col}` FROM `{table}` "
            f"WHERE `{col}` IS NOT NULL AND `{col}` REGEXP %s", (rx,))
        for r in cur.fetchall():
            s = str(r[col])
            pat = match_pattern(s)
            hits.append((str(r[pk])[:40], col, pat, s[:120]))
    return hits


def main():
    conn = db(OLD_DB)
    cur = conn.cursor()
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="C00000")
    head_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    summary = []
    for table, cols in TABLES.items():
        try:
            hits = scan(cur, table, cols) if len(cols) > 1 else []
        except Exception as e:
            print(f"  ⚠️ {table} 扫描失败: {e}")
            summary.append((table, -1))
            continue
        print(f"  {table}: 命中 {len(hits)} 条")
        summary.append((table, len(hits)))
        ws = wb.create_sheet(table[:31])
        ws.append(["主键", "命中字段", "命中特征", "命中内容样例"])
        for c in ws[1]:
            c.fill = header_fill
            c.font = head_font
        for row in hits:
            ws.append(list(row))
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = wrap
        for i, w in enumerate((42, 18, 24, 90), 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    ws = wb.create_sheet("汇总")
    ws.append(["表", "命中行数"])
    for t, n in summary:
        ws.append([t, n])
    for i, w in enumerate((24, 12), 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(OUT)
    conn.close()
    print(f"✅ 已生成 {OUT}")
    for t, n in summary:
        print(f"    {t}: {n}")


if __name__ == "__main__":
    main()
