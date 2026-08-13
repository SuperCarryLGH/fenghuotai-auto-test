#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成 影子表字段取值逻辑.xlsx：7 张影子表、每字段取值逻辑（源自迁移脚本 map_*）"""
import json
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SCHEMA = json.loads(Path("/tmp/shadow_schema.json").read_text())
OUT = "/Users/rs/Documents/影子表字段取值逻辑.xlsx"

# 每表取值逻辑：{字段: (取值逻辑, 源老字段, 备注)}
LOGIC = {
    "shadow_member_user": {
        "id": ("新用户=雪花gen_id；线上已存在用户=线上xlsx的member_user.id", "", ""),
        "mobile": ("老库手机号", "sys_user.phone", "按手机号去重取最新MAX(id)"),
        "password": ("不迁老密码，留空", "", ""),
        "status": ("老1→0；老{2,-1}/空→1", "sys_user.status", "映射"),
        "nickname": ("老昵称", "sys_user.nick_name", ""),
        "avatar": ("老默认头像→新默认(hy-recycle-mini.../default_avatar.webp)；空/None→NULL；其他原样", "sys_user.avatar", "开发确认"),
        "name": ("真实姓名，无实名留空", "sys_user.real_auth_name", ""),
        "sex": ("male/1/男→1；female/2/女→2；其余→0", "sys_user.sex", "对齐dev 0未知/1男/2女"),
        "station_id": ("站点ID", "sys_user.station_id", "int"),
        "company_id": ("公司ID", "sys_user.company_id", "int"),
        "channel": ("推广渠道", "sys_user.channel", ""),
        "promotion_site_id": ("推广站点", "sys_user.promotion_station_id", "int"),
        "promotion_activity_id": ("推广活动", "sys_user.promotion_activity_id", "int"),
        "promotion_platform": ("留空", "", "开发确认"),
        "promotion_channel": ("留空", "", "开发确认"),
        "is_promoter": ("固定0", "", ""),
        "level_id": ("固定1", "", "约定"),
        "point": ("固定0", "", ""),
        "experience": ("固定0", "", ""),
        "risk_level": ("固定0", "", ""),
        "risk_status": ("固定0", "", ""),
        "register_ip": ("固定空串", "", ""),
        "login_ip": ("固定空串", "", ""),
        "wx_transfer_openid": ("固定空串", "", ""),
        "ali_transfer_name": ("固定空串", "", ""),
        "ali_transfer_mobile": ("固定空串", "", ""),
        "ali_transfer_openid": ("固定空串", "", ""),
        "register_terminal": ("无源置NULL", "", ""),
        "login_date": ("无源置NULL", "", ""),
        "area_id": ("无源置NULL", "", ""),
        "birthday": ("无源置NULL", "", ""),
        "mark": ("无源置NULL", "", ""),
        "tag_ids": ("无源置NULL", "", ""),
        "group_id": ("无源置NULL", "", ""),
        "platform": ("无源置NULL", "", ""),
        "superior_promoter_id": ("无源置NULL", "", ""),
        "super_superior_promoter_id": ("无源置NULL", "", ""),
        "warehouse_id": ("无源置NULL", "", ""),
        "scene": ("无源置NULL", "", ""),
        "block_reason": ("无源置NULL", "", ""),
        "provider": ("无源置NULL", "", ""),
        "operation_center_id": ("后置回填：按老UUID匹配线上分拣中心映射", "backfill_operation_center_id.py", "13.7万命中"),
        "operation_center_uuid": ("老pay_station_id(UUID)", "sys_user.pay_station_id", "临时中转列，同步时排除"),
        "create_time": ("老创建时间", "sys_user.created_at", ""),
        "creator": ("固定'migrate'", "", "通用"),
        "update_time": ("迁移时刻", "", "通用"),
        "updater": ("固定'migrate'", "", "通用"),
        "deleted": ("固定0", "", "通用"),
        "tenant_id": ("固定1", "", "通用"),
    },
    "shadow_pay_wallet": {
        "id": ("新用户=雪花；线上已存在用户=继承线上wallet.id", "", "已存在=330个∩用户"),
        "user_id": ("by_account映射的uid", "wallet.owner_id(account_id)", "新=新雪花uid/已存在=线上uid"),
        "user_type": ("固定1(C端)", "", ""),
        "balance": ("新用户=老余额×100；已存在=线上余额+老余额×100（保留负）", "wallet.balance", "元→分"),
        "total_expense": ("固定0", "", ""),
        "total_recharge": ("固定0", "", ""),
        "freeze_price": ("固定0", "", ""),
        "sync_action": ("后置：已存在用户钱包='UPDATE'，其余='INSERT'", "mark_sync_action.py", "同步线上区分"),
        "create_time": ("老创建时间", "wallet.created_at", ""),
        "creator": ("固定'migrate'", "", "通用"),
        "update_time": ("迁移时刻", "", "通用"),
        "updater": ("固定'migrate'", "", "通用"),
        "deleted": ("固定0", "", "通用"),
        "tenant_id": ("固定1", "", "通用"),
    },
    "shadow_pay_wallet_transaction": {
        "id": ("雪花", "", "派生：老钱包余额>0各生成1条"),
        "wallet_id": ("对应影子钱包id", "", "已存在=继承线上wallet_id"),
        "biz_type": ("固定0(系统导入)", "", ""),
        "biz_id": ("固定空串", "", ""),
        "no": ("雪花，唯一流水号", "", ""),
        "title": ("固定'系统导入'", "", ""),
        "price": ("老库钱包余额（正=收入；负余额也生成负数流水；0金额不生成）", "wallet.balance", ""),
        "balance": ("交易后余额=dev现有+老库余额", "", "含已存在用户累加"),
        "trade_channel": ("无源置NULL", "", ""),
        "create_time": ("迁移时刻", "", ""),
        "creator": ("固定'migrate'", "", "通用"),
        "update_time": ("迁移时刻", "", "通用"),
        "updater": ("固定'migrate'", "", "通用"),
        "deleted": ("固定0", "", "通用"),
        "tenant_id": ("固定1", "", "通用"),
    },
    "shadow_member_address": {
        "id": ("雪花", "", "污染行整行跳过"),
        "user_id": ("by_account映射的uid", "user_address.account_id", ""),
        "name": ("老name，缺失→空串''(NOT NULL)，截断10", "user_address.name", "NOT NULL列不能用NULL"),
        "mobile": ("老phone，缺失→空串''(NOT NULL)，截断20", "user_address.phone", ""),
        "area_id": ("老district_code→int，缺失→0(NOT NULL)", "user_address.district_code", "bigint NOT NULL"),
        "province_code": ("老同名，缺失→NULL，截断32", "user_address.province_code", ""),
        "province": ("老同名，缺失→NULL，截断64", "user_address.province", ""),
        "city_code": ("老同名，缺失→NULL，截断32", "user_address.city_code", ""),
        "city": ("老同名，缺失→NULL，截断64", "user_address.city", ""),
        "district_code": ("老同名，缺失→NULL，截断32", "user_address.district_code", ""),
        "district": ("老同名，缺失→NULL，截断64", "user_address.district", ""),
        "detail_address": ("老address_detail，缺失→空串''(NOT NULL)，截断250", "user_address.address_detail", "NOT NULL列不能用NULL"),
        "default_status": ("老default?1:0", "user_address.default", ""),
        "door_plate": ("老house_no，缺失→NULL，截断100", "user_address.house_no", ""),
        "community_name": ("老同名，缺失→NULL，截断64", "user_address.community_name", ""),
        "community_code": ("老同名，缺失→NULL，截断32", "user_address.community_code", ""),
        "longitude": ("老longitude→decimal(可空)", "user_address.longitude", ""),
        "latitude": ("老latitude→decimal(可空)", "user_address.latitude", ""),
        "create_time": ("老创建时间", "user_address.created_at", ""),
        "update_time": ("老更新时间", "user_address.updated_at", ""),
        "creator": ("固定'migrate'", "", "通用"),
        "updater": ("固定'migrate'", "", "通用"),
        "deleted": ("固定0", "", "通用"),
        "tenant_id": ("固定1", "", "通用"),
    },
    "shadow_dist_promoter": {
        "id": ("雪花", "", ""),
        "user_id": ("老account_id→by_account uid（未命中跳过）", "promotion_promoter.account_id", ""),
        "apply_id": ("固定0", "", ""),
        "real_name": ("老实名", "sys_user.real_auth_name", ""),
        "id_card": ("老实名证件", "sys_user.real_auth_id", ""),
        "id_card_front": ("固定空串", "", ""),
        "id_card_back": ("固定空串", "", ""),
        "auth_status": ("有实名30/无实名10", "", ""),
        "operate_agreement_url": ("固定空串", "", ""),
        "operate_agreement_status": ("固定0", "", ""),
        "qrcode_wechat": ("无源置NULL", "", ""),
        "qrcode_alipay": ("无源置NULL", "", ""),
        "qrcode_douyin": ("无源置NULL", "", ""),
        "qrcode_kuaishou": ("无源置NULL", "", ""),
        "qrcode_normal": ("无源置NULL", "", ""),
        "parent_promoter_id": ("推广人账号的上一级（映射回填）", "sys_user.promoter_id", ""),
        "grand_promoter_id": ("推广人的上上级（映射回填）", "sys_user.promoter_agent_id", ""),
        "promoter_level": ("固定1", "", ""),
        "promoter_star": ("固定1", "", ""),
        "promoter_type": ("固定10", "", ""),
        "open_source": ("固定10", "", ""),
        "open_time": ("老审核通过时间", "promotion_promoter.approve_time", ""),
        "close_time": ("无源置NULL", "", ""),
        "status": ("固定1", "", ""),
        "team_id": ("固定0", "", ""),
        "first_promote_user_num": ("固定0", "", "统计字段"),
        "second_promote_user_num": ("固定0", "", "统计字段"),
        "first_order_count": ("固定0", "", "统计字段"),
        "second_order_count": ("固定0", "", "统计字段"),
        "first_order_amount": ("固定0", "", "统计字段"),
        "second_order_amount": ("固定0", "", "统计字段"),
        "first_order_weight": ("固定0", "", "统计字段"),
        "second_order_weight": ("固定0", "", "统计字段"),
        "first_order_complete_count": ("固定0", "", "统计字段"),
        "second_order_complete_count": ("固定0", "", "统计字段"),
        "real_name_auth_status": ("有实名1/无实名0", "", ""),
        "auth_remark": ("无源置NULL", "", ""),
        "auth_submit_time": ("无源置NULL", "", ""),
        "auth_check_time": ("无源置NULL", "", ""),
        "create_time": ("老创建时间", "promotion_promoter.created_at", ""),
        "creator": ("固定'migrate'", "", "通用"),
        "update_time": ("迁移时刻", "", "通用"),
        "updater": ("固定'migrate'", "", "通用"),
        "deleted": ("固定0", "", "通用"),
        "tenant_id": ("固定1", "", "通用"),
    },
    "shadow_dist_promoter_user_relation": {
        "id": ("雪花", "", ""),
        "promoter_id": ("老sys_user.promoter_id→新dist_promoter.id", "sys_user.promoter_id", ""),
        "promotor_user_id": ("推广人账号uid", "", ""),
        "user_id": ("绑定用户uid", "sys_user.account_id", ""),
        "user_name": ("老昵称", "sys_user.nick_name", ""),
        "avatar": ("老头像", "sys_user.avatar", ""),
        "parent_promoter_id": ("推广人的上一级", "sys_user.promoter_id(推广人)", ""),
        "promoter_type": ("固定1", "", ""),
        "team_id": ("无源置NULL", "", ""),
        "bind_source": ("固定10", "", ""),
        "bind_time": ("老创建时间", "sys_user.created_at", "实际绑定时间"),
        "status": ("固定1", "", ""),
        "remark": ("固定'用户注册时绑定推广关系'", "", ""),
        "create_time": ("迁移时刻", "", "技术字段=update_time"),
        "update_time": ("迁移时刻", "", "技术字段"),
        "creator": ("固定'migrate'", "", "通用"),
        "updater": ("固定'migrate'", "", "通用"),
        "deleted": ("固定0", "", "通用"),
        "tenant_id": ("固定1", "", "通用"),
    },
    "shadow_recycle_order": {
        "id": ("雪花", "", "开发逐项确认"),
        "order_no": ("雪花自生成（uk唯一）", "", "不映射老order_id"),
        "order_type": ("biz_mode：WeightClothes→0/SiteStationWeight→1/Exclusive→2/其他→0", "order.biz_mode", "在线0/面对面1/专属2"),
        "platform": ("老同名", "order.platform", ""),
        "provider": ("无源置NULL", "", ""),
        "channel": ("老有值即刷，无值→NULL", "order.channel", "开发确认"),
        "scene": ("无源置NULL", "", ""),
        "promoter_id": ("置NULL", "", "开发确认不处理"),
        "promotion_platform": ("置NULL", "", "开发确认留空"),
        "promotion_channel": ("置NULL", "", "开发确认留空"),
        "promotion_station_id": ("老同名", "order.promotion_station_id", ""),
        "user_id": ("老account_id→by_account uid", "order.account_id", ""),
        "user_phone": ("老同名", "order.user_phone", ""),
        "user_name": ("老同名，截断64", "order.user_name", ""),
        "province_code": ("老同名，截断32", "order.province_code", ""),
        "province": ("老同名，截断64", "order.province", ""),
        "city_code": ("老同名，截断32", "order.city_code", ""),
        "city": ("老同名，截断64", "order.city", ""),
        "district_code": ("老同名，截断32", "order.district_code", ""),
        "district": ("老同名，截断64", "order.district", ""),
        "appointment_date": ("老同名", "order.appointment_date", ""),
        "appointment_time_period": ("老同名", "order.appointment_time_period", ""),
        "appointment_week_str": ("老同名", "order.appointment_week_str", ""),
        "address_detail": ("老同名，截断255", "order.address_detail", ""),
        "appointment_photos": ("置NULL", "", "不迁老pics"),
        "express_type": ("老同名", "order.express_type", ""),
        "express_name": ("置NULL", "", ""),
        "express_order": ("老同名", "order.express_order", ""),
        "express_cost": ("老同名（decimal2位）", "order.express_cost", ""),
        "express_status": ("老同名", "order.express_status", ""),
        "express_meterage_weight": ("老同名（decimal2位）", "order.express_meterage_weight", ""),
        "pre_weight": ("老predict_weight（varchar去尾零）", "order.predict_weight", ""),
        "real_weight": ("线上=express_real_weight；线下=pay_kg（decimal3位）", "order.express_real_weight/order.pay_kg", "按biz_mode分支"),
        "package_num": ("老同名→int，默认0", "order.package_num", ""),
        "total_price": ("都写老pay_money（元，2位）", "order.pay_money", "开发确认"),
        "pay_price": ("都写老pay_money（元，2位）", "order.pay_money", "开发确认"),
        "pay_type": ("固定2（钱包余额）", "", "开发确认统一"),
        "settlement_type": ("settle_type：ExpressWeight→1/FactoryQa→2/ActivityRewardDonate→1", "order.operation_center_settle_type", "1回收/2到场"),
        "cancel_time": ("置NULL", "", ""),
        "status": ("官方表映射：老1→10/3→10/4→10/5→20/6·10~60→30/999→30/-1·-2·-10→50/-3·2·-20→NULL", "order.status", "order-recycle_order.xlsx"),
        "clear_status": ("固定0", "", "NOT NULL"),
        "sub_status": ("官方表映射：1→11/3→21/4→22/5→23/已完成(6·10~60)→32/999→32/-1→51/-2→52/-10→55/未覆盖→NULL", "order.status", "order-recycle_order.xlsx"),
        "inspect_status": ("官方表：6·10~60→20，其余NULL", "order.status", "order-recycle_order.xlsx"),
        "cancel_type": ("官方表：-1→0/-2→1/-10→2，其余NULL", "order.status", "order-recycle_order.xlsx"),
        "activity_id": ("老同名→int", "order.activity_id", ""),
        "lat": ("老lat=纬度（实测）", "order.lat", "老表注释写反，数据正确"),
        "lon": ("老lon=经度（实测）", "order.lon", "老表注释写反，数据正确"),
        "receive_time": ("老同名（1970占位→NULL）", "order.service_station_receive_order_time", ""),
        "recycle_begin_time": ("老start_to_door_time（1970占位→NULL）", "order.service_station_start_to_door_time", ""),
        "recycle_end_time": ("老recycle_completed_time（1970占位→NULL）", "order.service_station_recycle_completed_time", ""),
        "pay_time": ("=recycle_end_time（service_station_recycle_completed_time，1970→NULL）", "order.service_station_recycle_completed_time", "开发确认"),
        "inspect_time": ("老finish_time（1970占位→NULL）", "order.operation_center_finish_time", ""),
        "settlement_status": ("官方表：6·10~60→30，其余NULL", "order.status", "order-recycle_order.xlsx"),
        "weight_time": ("复用老recycle_completed_time（1970占位→NULL）", "order.service_station_recycle_completed_time", "开发确认复用"),
        "recycler_user_id": ("线上=老接收订单账号；线下=NULL", "order.service_station_receive_order_account_id", "按biz_mode分支"),
        "recycler_user_name": ("置NULL", "", ""),
        "recycler_user_phone": ("线上=老接单三方电话；线下=NULL", "order.service_station_receive_tp_phone", "按biz_mode分支"),
        "station_id": ("置NULL", "", "开发确认不管"),
        "warehouse_id": ("置NULL", "", "开发确认不管"),
        "operation_center_id": ("置NULL（后置刷分拣中心）", "", "待后置脚本"),
        "company_id": ("置NULL", "", ""),
        "express_status_desc": ("置NULL", "", ""),
        "door_plate": ("老house_num，截断100", "order.house_num", ""),
        "community_name": ("老同名，截断64", "order.community_name", ""),
        "community_code": ("老同名，截断32", "order.community_code", ""),
        "address_longitude": ("置NULL", "", ""),
        "address_latitude": ("置NULL", "", ""),
        "detail_address": ("置NULL", "", "开发确认留空/不处理"),
        "express_emp_code": ("线上=老接收订单账号；线下=NULL", "order.service_station_receive_order_account_id", "按biz_mode分支"),
        "express_emp_phone": ("线上=老接单三方电话；线下=NULL", "order.service_station_receive_tp_phone", "按biz_mode分支"),
        "express_net_code": ("线上=老服务站id；线下=NULL", "order.service_station_id", "按biz_mode分支"),
        "volume": ("置NULL", "", ""),
        "cancel_reason": ("置NULL", "", ""),
        "address_id": ("固定0", "", "NOT NULL"),
        "third_order_no": ("置NULL", "", ""),
        "create_time": ("老创建时间", "order.created_at", ""),
        "creator": ("固定'migrate'", "", "通用"),
        "update_time": ("迁移时刻", "", "通用"),
        "updater": ("固定'migrate'", "", "通用"),
        "deleted": ("固定0", "", "通用"),
        "tenant_id": ("固定1", "", "通用"),
    },
    "shadow_recycle_order_item": {
        "id": ("雪花", "", ""),
        "order_id": ("新recycle_order.id（订单+明细同脚本同批生成，位置配对）", "", "订单过滤后1:1"),
        "item_id": ("固定2047530778823024642", "", "ERP统货产品id"),
        "item_code": ("订单package_code位置码（第i明细↔第i包裹码，截断64）", "order.package_code切值", "与package_no逐条一致"),
        "item_name": ("固定'统货'", "", ""),
        "item_unit": ("固定'KG'", "", ""),
        "item_pic": ("统一NULL", "", "开发确认"),
        "price": ("老recycle_price（元，2位）", "order_product.recycle_price", ""),
        "weight": ("老num（kg，3位）", "order_product.num", ""),
        "total_price": ("recycle_price×num（元，2位）", "order_product.recycle_price*num", ""),
        "create_time": ("老创建时间", "order_product.created_at", ""),
        "update_time": ("老更新时间", "order_product.updated_at", ""),
        "creator": ("固定'migrate'", "", "通用"),
        "updater": ("固定'migrate'", "", "通用"),
        "deleted": ("固定0", "", "通用"),
        "tenant_id": ("固定1", "", "通用"),
    },
    "shadow_recycle_package_item": {
        "id": ("雪花", "", "纯dev派生：recycle_order_item→package"),        "package_no": ("=recycle_order_item.item_code（订单package_code位置码）", "item.item_code", "跨订单唯一，与明细一致"),
        "recycle_order_id": ("=item.order_id", "item.order_id", ""),
        "item_id": ("=item.item_id", "item.item_id", "固定2047530778823024642"),
        "item_name": ("=item.item_name", "item.item_name", "'统货'"),
        "item_unit": ("=item.item_unit", "item.item_unit", "'KG'"),
        "item_pic_url": ("统一NULL", "", "开发确认"),
        "recycle_price": ("=item.price", "item.price", ""),
        "recycle_weight": ("=item.weight", "item.weight", ""),
        "recycle_total_price": ("=item.total_price", "item.total_price", ""),
        "recycle_time": ("=order.create_time", "order.create_time", ""),
        "recycle_recive_time": ("=order.receive_time", "order.receive_time", ""),
        "recycle_pay_time": ("=order.pay_time（NULL）", "order.pay_time", ""),
        "package_status": ("固定101", "", "开发代码"),
        "settle_status": ("订单settlement_status=30→1(已结算)，否则0(未结算)", "order.settlement_status", "动态"),
        "stock_status": ("固定0（未入库）", "", ""),
        "inspect_status": ("订单inspect_status=20→30(审核通过)，否则10(待质检)", "order.inspect_status", "动态"),
        "transfer_status": ("固定0", "", "NOT NULL"),
        "creator": ("固定'migrate'", "", "通用"),
        "create_time": ("=order.create_time", "order.create_time", ""),
        "updater": ("固定'migrate'", "", "通用"),
        "update_time": ("迁移时刻", "", "通用"),
        "deleted": ("固定0", "", "通用"),
        "tenant_id": ("固定1", "", "通用"),
    },
    "shadow_dist_commission_account": {
        "id": ("雪花", "", "纯dev派生：dist_promoter→佣金账户"),
        "account_type": ("固定1", "", "开发确认"),
        "account_id": ("=dist_promoter.id", "dist_promoter.id", "推广员佣金账户，天然1:1"),
        "balance": ("固定0", "", ""),
        "total_income": ("固定0", "", ""),
        "total_expend": ("固定0", "", ""),
        "total_freeze": ("固定0", "", ""),
        "remark": ("置NULL", "", ""),
        "creator": ("固定'migrate'", "", "通用"),
        "create_time": ("=dist_promoter.create_time", "dist_promoter.create_time", ""),
        "updater": ("固定'migrate'", "", "通用"),
        "update_time": ("=dist_promoter.update_time", "dist_promoter.update_time", ""),
        "deleted": ("固定0", "", "通用"),
        "tenant_id": ("固定1", "", "通用"),
    },
}

HEADER_NOTE = {
    "shadow_member_user": "源:sys_user(role_id=5,手机号去重取最新) | 约142.2万行 | 已存在330个线上用户继承线上id",
    "shadow_pay_wallet": "源:wallet(owner_id=account_id) | 约142.2万行 | 余额元→分×100 | 已存在用户累加+继承线上wallet_id",
    "shadow_pay_wallet_transaction": "派生:老钱包余额>0各生成1条 | 约3.7万行 | biz_type=0系统导入",
    "shadow_member_address": "源:user_address | 约19万行(污染688被过滤) | 字符串缺失→NULL/超长截断",
    "shadow_dist_promoter": "源:promotion_promoter | 约42.6万行 | 统计/等级/身份固定值",
    "shadow_dist_promoter_user_relation": "源:sys_user.promoter_id | 约118万行 | bind_time=老created_at",
    "shadow_recycle_order": "源:order | 可迁约34.6万行(过滤18.2%=空码17.9%+不匹配0.4%) | 状态按官方表",
    "shadow_recycle_order_item": "并入migrate_order.py单脚本 | 约40.4万行 | item_code=订单package_code位置码 | 与包裹逐条一致",
    "shadow_recycle_package_item": "纯dev派生 | 每明细1包裹 | package_no=item_code | 状态101/0/0/10/0",
    "shadow_dist_commission_account": "纯dev派生 | 每推广员1佣金账户 | account_id=dist_promoter.id | 金额全0",
}


def main():
    # 兜底：LOGIC 未显式写明的字段 → 统一标注"无源置NULL"
    for table in SCHEMA:
        logic = LOGIC.setdefault(table, {})
        for col in SCHEMA[table]:
            logic.setdefault(col["f"], ("无源置NULL", "", ""))
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="305496")
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    head_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    for table in SCHEMA:
        ws = wb.create_sheet(table.replace("shadow_", "", 1))
        # 顶部说明
        ws.append([HEADER_NOTE[table]])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        for c in ws[1]:
            c.fill = note_fill
            c.alignment = wrap
        ws.append(["序号", "字段名", "类型", "可空", "取值逻辑", "源老字段", "备注"])
        for c in ws[2]:
            c.fill = header_fill
            c.font = head_font
        for i, col in enumerate(SCHEMA[table], 1):
            f = col["f"]
            logic, src, note = LOGIC.get(table, {}).get(f, ("", "", "")) or ("", "", "")
            ws.append([i, f, col["type"], "是" if col["null"] == "YES" else "否", logic, src, note])
        # 列宽
        widths = [6, 26, 18, 6, 60, 30, 20]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w
        for row in ws.iter_rows(min_row=3):
            for c in row:
                c.alignment = wrap
        ws.freeze_panes = "A3"
    wb.save(OUT)
    print(f"✅ 已生成 {OUT}")
    for t in SCHEMA:
        n = len(SCHEMA[t])
        missing = [c for c in SCHEMA[t] if c["f"] not in LOGIC.get(t, {})]
        print(f"  {t}: {n}列" + (f" ❌缺逻辑: {missing}" if missing else " ✓ 逻辑齐全"))


if __name__ == "__main__":
    main()
