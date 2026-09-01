"""
BD签约全流程 - Step 2: 创建线索
"""
import sys
import os
import json
import time
import argparse
from openpyxl import load_workbook

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, PROJECT_ROOT)

from config import CENTER_MAP, PROVINCE_CODE_MAP, COMMON_HEADERS, SUPER_ADMIN
import pymysql
import requests


def get_db():
    """获取数据库连接"""
    return pymysql.connect(
        host='rm-bp1kmprsfdog024fsro.mysql.rds.aliyuncs.com',
        port=3306, user='sf_fht_dev', password='8HUvyZf6X&FNR%5',
        database='fht_yhs', charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor, connect_timeout=5
    )


def read_excel(excel_path, bd_phones=None):
    """读取Excel数据
    
    参数:
        excel_path: 站点Excel路径
        bd_phones: BD手机号集合（可选，用于过滤）
    
    Excel列映射（从0开始）:
    0: 序号
    1: 网点类型
    2: 推广平台
    3: 推广渠道
    4: 站点类型
    5: 运营模式
    6: 回收单价
    7: 支付类型
    8: 清运方式
    9: 呼叫方式
    10: 起运重量
    11: 收益结算方式
    12: 收益到账
    13: 清运结算价格
    14: 清运收益
    15: 开票方式
    16: 提现方式
    17: 下单物料
    18: 宣传物料
    19: 培训物料
    20: 站点名称
    21: 负责人
    22: 电话
    23: 所属省份
    24: 所属城市
    25: 所属区县
    26: 所属街道/村/社区
    27: 详细地址
    28: 站点关联的BD名称
    29: 关联BD的手机号
    """
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or not row[20] or not row[22]:  # 跳过空行和无站点名称/电话的行
            continue
        
        phone = str(row[22]).strip() if row[22] else ""
        if len(phone) == 11 and phone.startswith('1'):
            province = str(row[23]).strip() if row[23] else ""
            province_code = PROVINCE_CODE_MAP.get(province, "")
            if not province_code:
                for key, code in PROVINCE_CODE_MAP.items():
                    if key in province or province in key:
                        province_code = code
                        break
            
            bd_phone = str(row[29]).strip() if row[29] else ""
            
            # 如果指定了BD手机号集合，过滤不在集合中的
            if bd_phones is not None and bd_phone not in bd_phones:
                continue
            
            data.append({
                "row": row_idx,
                "station_name": str(row[20]).strip() if row[20] else "",
                "contact_name": str(row[21]).strip() if row[21] else "",
                "contact_phone": phone,
                "province": province,
                "province_code": province_code or "330000",
                "city": str(row[24]).strip() if row[24] else "",
                "district": str(row[25]).strip() if row[25] else "",
                "detail_address": str(row[27]).strip() if row[27] else "",
                "station_type": 1 if str(row[4]).strip() == "门店" else 2 if str(row[4]).strip() == "回收机" else 3,
                "industry": str(row[5]).strip() if row[5] else "默认",
                "bd_name": str(row[28]).strip() if row[28] else "",
                "bd_phone": bd_phone,
                "bd_center": str(row[11]).strip() if row[11] else "",
            })
    
    wb.close()
    return data


def check_existing_clues(phones):
    """检查线索是否已存在"""
    db = get_db()
    with db.cursor() as cur:
        placeholders = ','.join(['%s'] * len(phones))
        cur.execute(f'SELECT contact_phone FROM station_clue WHERE contact_phone IN ({placeholders}) AND deleted = 0', phones)
        existing = set(row['contact_phone'] for row in cur.fetchall())
    db.close()
    return existing


def get_center_id(center_name):
    """获取分拣中心ID"""
    db = get_db()
    with db.cursor() as cur:
        cur.execute('SELECT id FROM station WHERE name = %s AND deleted = 0', (center_name,))
        row = cur.fetchone()
    db.close()
    return row['id'] if row else None


def get_super_admin_token():
    """获取超级管理员token"""
    resp = requests.post(
        f"https://api-fht-dev.hengyishou.com/admin-api/system/auth/login",
        json=SUPER_ADMIN,
        headers=COMMON_HEADERS,
        timeout=10,
    )
    data = resp.json()
    if data.get("code") == 0:
        return data["data"]["accessToken"]
    return None


def create_clue(token, clue_data, center_id):
    """创建单个线索"""
    headers = COMMON_HEADERS.copy()
    headers["Authorization"] = f"Bearer {token}"
    
    detail_address = clue_data['detail_address']
    if not detail_address:
        detail_address = f"{clue_data['province']}{clue_data['city']}{clue_data['district']}"
    
    # 使用地理编码获取 cityCode、districtCode、经纬度
    from geocoder import geocode_address
    geo_result = geocode_address(
        detail_address,
        clue_data.get('province', ''),
        clue_data.get('city', ''),
        clue_data.get('district', ''),
    )
    
    body = {
        "poolType": 0,  # 公海线索
        "clueName": clue_data['station_name'],
        "contactName": clue_data['contact_name'],
        "contactPhone": clue_data['contact_phone'],
        "stationType": clue_data['station_type'],
        "detailAddress": detail_address,
        "province": clue_data['province'] or "浙江省",
        "provinceCode": geo_result.get("province_code") or clue_data.get('province_code', "330000"),
        "city": clue_data['city'] or "杭州市",
        "cityCode": geo_result.get("city_code") or "331000",
        "district": clue_data['district'] or "滨江区",
        "districtCode": geo_result.get("district_code") or "330108",
        "lat": geo_result.get("lat", 0),
        "lon": geo_result.get("lon", 0),
        "belongCenterId": center_id,
    }
    
    try:
        resp = requests.post(
            "https://api-fht-dev.hengyishou.com/admin-api/recycle/station-clue/create",
            json=body,
            headers=headers,
            timeout=10,
        )
        return resp.json()
    except Exception as e:
        return {"code": -1, "msg": str(e)}


def create_clues(excel_path, test_count=0, bd_phones=None):
    """批量创建线索
    
    参数:
        excel_path: 站点Excel路径
        test_count: 测试数量
        bd_phones: BD手机号集合（可选，用于过滤）
    """
    print("=" * 60)
    print("Step 2: 创建线索")
    print("=" * 60)
    
    # 1. 读取Excel（如果指定了bd_phones，则只返回BD在列表中的数据）
    clue_list = read_excel(excel_path, bd_phones)
    print(f"\n可处理的站点: {len(clue_list)} 个")
    
    if test_count > 0:
        clue_list = clue_list[:test_count]
        print(f"测试模式: 只处理前 {test_count} 个")
    
    # 2. 检查已存在的线索
    phones = [c['contact_phone'] for c in clue_list]
    existing = check_existing_clues(phones)
    print(f"已存在: {len(existing)} 个")
    
    # 3. 获取超级管理员token
    token = get_super_admin_token()
    if not token:
        print("❌ 获取超级管理员token失败")
        return {"total": len(clue_list), "existing": 0, "created": 0, "failed": len(clue_list)}
    
    print("✅ 超级管理员登录成功")
    
    # 4. 创建线索
    created = 0
    skipped = 0
    failed = 0
    
    for i, clue in enumerate(clue_list):
        if clue['contact_phone'] in existing:
            skipped += 1
            continue
        
        # 获取分拣中心ID
        center_id = get_center_id(clue['bd_center'])
        if not center_id:
            print(f"  ⚠️ [{i+1}] 未找到分拣中心: {clue['bd_center']}")
            failed += 1
            continue
        
        result = create_clue(token, clue, center_id)
        if result.get("code") == 0:
            created += 1
            if created % 10 == 0:
                print(f"  进度: {i+1}/{len(clue_list)}, 创建: {created}")
        else:
            failed += 1
            if failed <= 5:
                print(f"  ❌ [{i+1}] {clue['station_name']}: {result.get('msg')}")
        
        time.sleep(0.3)
    
    # 5. 输出汇总
    print(f"\n{'='*60}")
    print("线索创建汇总")
    print(f"{'='*60}")
    print(f"总计: {len(clue_list)}")
    print(f"已存在: {skipped}")
    print(f"新创建: {created}")
    print(f"失败: {failed}")
    
    return {"total": len(clue_list), "existing": skipped, "created": created, "failed": failed}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="创建线索")
    parser.add_argument("--excel", required=True, help="Excel文件路径")
    parser.add_argument("--test", type=int, default=0, help="测试模式：只处理前N个")
    args = parser.parse_args()
    
    if not os.path.exists(args.excel):
        print(f"❌ 文件不存在: {args.excel}")
        sys.exit(1)
    
    create_clues(args.excel, args.test)
