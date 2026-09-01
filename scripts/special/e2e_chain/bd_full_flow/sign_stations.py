"""
BD签约全流程 - Step 3: 签约站点
"""
import sys
import os
import json
import time
import argparse
from openpyxl import load_workbook

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, PROJECT_ROOT)

from config import INDUSTRY_MAP, PROVINCE_CODE_MAP, COMMON_HEADERS
import pymysql
import requests


def get_db():
    """获取数据库连接"""
    return pymysql.connect(
        host='rm-bp1kmprsfdog024fsro.mysql.rds.aliyuncs.com',
        port=3306, user='sf_fht_dev', password='8HUvyZf6X&FNR%5',
        database='fht_yhs', charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor, connect_timeout=5
    )


def read_excel(excel_path, bd_phones=None):
    """读取Excel数据
    
    参数:
        excel_path: 站点Excel路径
        bd_phones: BD手机号集合（可选，用于过滤）
    """
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or not row[20] or not row[22]:
            continue
        
        phone = str(row[22]).strip() if row[22] else ""
        if len(phone) == 11 and phone.startswith('1'):
            industry_name = str(row[5]).strip() if row[5] else "默认"
            industry_info = INDUSTRY_MAP.get(industry_name, INDUSTRY_MAP["默认"])
            province = str(row[23]).strip() if row[23] else ""
            province_code = ""
            for key, code in PROVINCE_CODE_MAP.items():
                if key in province or province in key:
                    province_code = code
                    break
            
            bd_phone = str(row[29]).strip() if row[29] else ""
            
            # 如果指定了BD手机号集合，过滤不在集合中的
            if bd_phones is not None and bd_phone not in bd_phones:
                continue
            
            data.append({
                "row": row_idx,
                "station_name": str(row[20]).strip() if row[20] else "",
                "contact_name": str(row[21]).strip() if row[21] else "",
                "contact_phone": phone,
                "industry": industry_info["value"],
                "industry_key": industry_info["key"],
                "operation_mode": "自营",
                "recycle_price": 0.5,
                "clear_price": 1.0,
                "province": province or "浙江省",
                "province_code": province_code or "330000",
                "city": str(row[24]).strip() if row[24] else "杭州市",
                "district": str(row[25]).strip() if row[25] else "滨江区",
                "street": str(row[26]).strip() if row[26] else "测试街道",
                "detail_address": str(row[27]).strip() if row[27] else "测试地址",
                "bd_name": str(row[28]).strip() if row[28] else "",
                "bd_phone": bd_phone,
            })
    
    wb.close()
    return data


def find_clue(phone):
    """查询线索"""
    db = get_db()
    with db.cursor() as cur:
        cur.execute('''
            SELECT id, status, pool_type, receive_user_id, belong_center_id
            FROM station_clue 
            WHERE contact_phone = %s AND deleted = 0
        ''', (phone,))
        row = cur.fetchone()
    db.close()
    return row


def get_bd_info(phone):
    """获取BD信息"""
    db = get_db()
    with db.cursor() as cur:
        cur.execute('SELECT id, nickname, operation_center_id FROM system_users WHERE mobile = %s', (phone,))
        row = cur.fetchone()
    db.close()
    return row


def login_bd(phone):
    """BD登录"""
    try:
        resp = requests.post(
            f"https://api-fht-dev.hengyishou.com/admin-api/system/auth/sms-login",
            json={"mobile": phone, "code": "9999"},
            headers=COMMON_HEADERS,
            timeout=10,
        )
        data = resp.json()
        if data.get("code") == 0:
            return data["data"]["accessToken"]
        return None
    except:
        return None


def claim_clue(token, clue_id):
    """领取线索"""
    headers = COMMON_HEADERS.copy()
    headers["Authorization"] = f"Bearer {token}"
    try:
        resp = requests.post(
            f"https://api-fht-dev.hengyishou.com/admin-api/recycle/station-clue/claim?id={clue_id}",
            headers=headers,
            timeout=10,
        )
        return resp.json()
    except Exception as e:
        return {"code": -1, "msg": str(e)}


def sign_station(token, clue_id, data):
    """签约站点"""
    headers = COMMON_HEADERS.copy()
    headers["Authorization"] = f"Bearer {token}"
    
    # 使用地理编码获取经纬度
    from geocoder import geocode_address
    geo_result = geocode_address(
        data.get("detail_address", ""),
        data.get("province", ""),
        data.get("city", ""),
        data.get("district", ""),
    )
    
    body = {
        "clueId": clue_id,
        "type": 31,
        "industry": data["industry"],
        "industryKey": data["industry_key"],
        "operationMode": data["operation_mode"],
        "recyclePrice": data["recycle_price"],
        "clearPrice": data["clear_price"],
        "stationName": data["station_name"],
        "contactName": data["contact_name"],
        "contactPhone": data["contact_phone"],
        "province": data["province"],
        "provinceCode": geo_result.get("province_code") or data["province_code"],
        "city": data["city"],
        "district": data["district"],
        "street": data["street"],
        "detailAddress": data["detail_address"],
        "lat": geo_result.get("lat", 0),
        "lon": geo_result.get("lon", 0),
        "paymentType": 10,
        "cleartMode": 10,
        "callMode": 10,
        "minWeight": 50,
        "settlementType": 10,
        "incomeType": 1,
        "invoiceType": 1,
        "withdrawType": 1,
    }
    
    try:
        resp = requests.post(
            "https://api-fht-dev.hengyishou.com/admin-api/recycle/station-clue/sign-submit",
            json=body,
            headers=headers,
            timeout=30,
        )
        return resp.json()
    except Exception as e:
        return {"code": -1, "msg": str(e)}


def load_bd_phones(bd_excel_path):
    """从BD Excel加载BD手机号"""
    bd_phones = set()
    try:
        wb = load_workbook(bd_excel_path, read_only=True)
        ws = wb.active
        current_center = ""
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and str(row[1]).strip():
                current_center = str(row[1]).strip()
            phone = str(row[6]).strip() if row[6] else ""
            if phone and len(phone) == 11 and phone.startswith('1'):
                bd_phones.add(phone)
        wb.close()
    except Exception as e:
        print(f"⚠️ 读取BD Excel失败: {e}")
    return bd_phones


def sign_stations(excel_path, bd_excel_path=None, test_count=0):
    """批量签约站点
    
    参数:
        excel_path: 站点Excel路径
        bd_excel_path: BD Excel路径（可选，用于过滤）
        test_count: 测试数量
    """
    print("=" * 60)
    print("Step 3: 签约站点")
    print("=" * 60)
    
    # 加载BD手机号
    bd_phones = None
    if bd_excel_path:
        bd_phones = load_bd_phones(bd_excel_path)
        print(f"\nBD Excel中的BD: {len(bd_phones)} 个")
    
    # 1. 读取Excel
    station_list = read_excel(excel_path, bd_phones)
    print(f"可处理的站点: {len(station_list)} 个")
    
    if test_count > 0:
        station_list = station_list[:test_count]
        print(f"测试模式: 只处理前 {test_count} 个")
    
    # 2. 签约流程
    success = 0
    fail = 0
    skip = 0
    tokens = {}
    results = {"success": [], "failed": []}
    
    for i, station in enumerate(station_list):
        phone = station['contact_phone']
        bd_phone = station['bd_phone']
        
        # 查询线索
        clue = find_clue(phone)
        if not clue:
            fail += 1
            results["failed"].append({"phone": phone, "station": station['station_name'], "error": "未找到线索"})
            continue
        
        if clue['status'] == 41:
            skip += 1
            continue
        
        if clue['status'] not in (10, 20):
            skip += 1
            continue
        
        # 获取BD信息
        bd_info = get_bd_info(bd_phone)
        if not bd_info:
            fail += 1
            results["failed"].append({"phone": phone, "station": station['station_name'], "error": "未找到BD账号"})
            continue
        
        # 登录BD
        if bd_phone not in tokens:
            token = login_bd(bd_phone)
            if not token:
                fail += 1
                results["failed"].append({"phone": phone, "station": station['station_name'], "error": "登录失败"})
                continue
            tokens[bd_phone] = token
        
        # claim线索（如果是status=10）
        if clue['status'] == 10:
            claim_result = claim_clue(tokens[bd_phone], clue['id'])
            if claim_result.get('code') != 0:
                fail += 1
                results["failed"].append({"phone": phone, "station": station['station_name'], "error": f"领取失败: {claim_result.get('msg')}"})
                continue
        
        # sign签约
        sign_result = sign_station(tokens[bd_phone], clue['id'], station)
        if sign_result.get('code') == 0:
            sign_data = sign_result.get('data', {})
            success += 1
            results["success"].append({
                "phone": phone,
                "station": station['station_name'],
                "clueId": clue['id'],
                "signId": sign_data.get('signId'),
                "stationId": sign_data.get('stationId'),
            })
            if success % 10 == 0:
                print(f"  进度: {i+1}/{len(station_list)}, 成功: {success}")
        else:
            fail += 1
            results["failed"].append({"phone": phone, "station": station['station_name'], "error": sign_result.get('msg')})
            if fail <= 5:
                print(f"  ❌ [{i+1}] {station['station_name']}: {sign_result.get('msg')}")
        
        time.sleep(0.5)
    
    # 3. 输出汇总
    print(f"\n{'='*60}")
    print("签约汇总")
    print(f"{'='*60}")
    print(f"总计: {len(station_list)}")
    print(f"成功: {success}")
    print(f"失败: {fail}")
    print(f"跳过: {skip}")
    
    # 4. 保存结果
    result_file = os.path.join(os.path.dirname(__file__), "sign_result.json")
    with open(result_file, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\n结果已保存: {result_file}")
    
    return {"total": len(station_list), "success": success, "failed": fail, "skip": skip}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="签约站点")
    parser.add_argument("--excel", required=True, help="Excel文件路径")
    parser.add_argument("--test", type=int, default=0, help="测试模式：只处理前N个")
    args = parser.parse_args()
    
    if not os.path.exists(args.excel):
        print(f"❌ 文件不存在: {args.excel}")
        sys.exit(1)
    
    sign_stations(args.excel, args.test)
