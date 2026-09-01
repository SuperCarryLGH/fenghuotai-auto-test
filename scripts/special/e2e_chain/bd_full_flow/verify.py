"""
BD签约全流程 - Step 4: 验证结果
"""
import sys
import os
import json
import argparse
from openpyxl import load_workbook

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, PROJECT_ROOT)

import pymysql


def get_db():
    """获取数据库连接"""
    return pymysql.connect(
        host='rm-bp1kmprsfdog024fsro.mysql.rds.aliyuncs.com',
        port=3306, user='sf_fht_dev', password='8HUvyZf6X&FNR%5',
        database='fht_yhs', charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor, connect_timeout=5
    )


def read_excel(excel_path):
    """读取Excel数据"""
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    phones = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row and row[2]:
            phone = str(row[2]).strip() if row[2] else ""
            if len(phone) == 11 and phone.startswith('1'):
                phones.append(phone)
    
    wb.close()
    return phones


def verify_results(excel_path):
    """验证签约结果"""
    print("=" * 60)
    print("Step 4: 验证结果")
    print("=" * 60)
    
    # 1. 读取Excel
    phones = read_excel(excel_path)
    print(f"\nExcel中的站点: {len(phones)} 个")
    
    db = get_db()
    
    # 2. 查询线索状态
    with db.cursor() as cur:
        placeholders = ','.join(['%s'] * len(phones))
        cur.execute(f'''
            SELECT status, COUNT(*) as cnt
            FROM station_clue 
            WHERE contact_phone IN ({placeholders}) AND deleted = 0
            GROUP BY status
        ''', phones)
        clue_status = {row['status']: row['cnt'] for row in cur.fetchall()}
    
    print(f"\n线索状态分布:")
    status_names = {10: "待处理", 20: "已领取", 41: "已签约"}
    for status, cnt in clue_status.items():
        print(f"  {status} ({status_names.get(status, '未知')}): {cnt}")
    
    # 3. 查询签约记录
    with db.cursor() as cur:
        cur.execute('''
            SELECT COUNT(*) as cnt
            FROM station_sign 
            WHERE deleted = 0
        ''')
        total_signs = cur.fetchone()['cnt']
    
    print(f"\n签约记录总数: {total_signs}")
    
    # 4. 查询站点数量
    with db.cursor() as cur:
        cur.execute('''
            SELECT COUNT(*) as cnt
            FROM station 
            WHERE deleted = 0
        ''')
        total_stations = cur.fetchone()['cnt']
    
    print(f"站点总数: {total_stations}")
    
    db.close()
    
    # 5. 汇总
    print(f"\n{'='*60}")
    print("验证汇总")
    print(f"{'='*60}")
    print(f"Excel站点数: {len(phones)}")
    print(f"已签约线索: {clue_status.get(41, 0)}")
    print(f"签约记录数: {total_signs}")
    print(f"站点总数: {total_stations}")
    
    return {
        "excel_count": len(phones),
        "signed_clues": clue_status.get(41, 0),
        "total_signs": total_signs,
        "total_stations": total_stations,
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="验证签约结果")
    parser.add_argument("--excel", required=True, help="Excel文件路径")
    args = parser.parse_args()
    
    if not os.path.exists(args.excel):
        print(f"❌ 文件不存在: {args.excel}")
        sys.exit(1)
    
    verify_results(args.excel)
