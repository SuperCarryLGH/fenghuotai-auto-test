"""
BD签约全流程 - Step 1: 创建BD账号
"""
import sys
import os
import json
import hashlib
import time
import argparse
from openpyxl import load_workbook

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, PROJECT_ROOT)

from config import CENTER_MAP, COMMON_HEADERS, SUPER_ADMIN, EXCEL_COLUMN_MAP
import pymysql
import requests


def get_db():
    """获取数据库连接"""
    return pymysql.connect(
        host='rm-bp1kmprsfdog024fsro.mysql.rds.aliyuncs.com',
        port=3306, user='sf_fht_dev', password='8HUvyZf6X&FNR%5',
        database='fht_yhs', charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor, connect_timeout=5
    )


def read_excel(excel_path):
    """读取Excel数据"""
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if row and row[0] and row[9] and row[10]:  # 需要BD姓名和手机号
            bd_phone = str(row[10]).strip() if row[10] else ""
            bd_name = str(row[9]).strip() if row[9] else ""
            bd_center = str(row[11]).strip() if row[11] else ""
            
            if len(bd_phone) == 11 and bd_phone.startswith('1'):
                data.append({
                    "row": row_idx + 1,
                    "bd_name": bd_name,
                    "bd_phone": bd_phone,
                    "bd_center": bd_center,
                })
    
    wb.close()
    
    # 按BD手机号去重
    unique_bds = {}
    for item in data:
        if item["bd_phone"] not in unique_bds:
            unique_bds[item["bd_phone"]] = item
    
    return list(unique_bds.values())


def check_existing_bd(bd_phones):
    """检查BD账号是否已存在"""
    db = get_db()
    with db.cursor() as cur:
        placeholders = ','.join(['%s'] * len(bd_phones))
        cur.execute(f'SELECT mobile FROM system_users WHERE mobile IN ({placeholders})', bd_phones)
        existing = set(row['mobile'] for row in cur.fetchall())
    db.close()
    return existing


def create_bd_account(bd_data, dept_id=None, warehouse_id=None):
    """创建单个BD账号"""
    db = get_db()
    try:
        with db.cursor() as cur:
            # 获取参考数据
            if not dept_id or not warehouse_id:
                cur.execute('SELECT dept_id, warehouse_id FROM system_users WHERE mobile = "18105736340"')
                ref = cur.fetchone()
                dept_id = dept_id or ref['dept_id']
                warehouse_id = warehouse_id or ref['warehouse_id']
            
            # 获取 operation_center_id
            center_name = bd_data['bd_center']
            operation_center_id = CENTER_MAP.get(center_name)
            if not operation_center_id:
                print(f"    ⚠️ 未找到分拣中心映射: {center_name}")
                return False
            
            # 创建账号
            password_hash = hashlib.sha256("123456".encode()).hexdigest()
            cur.execute('''
                INSERT INTO system_users (
                    username, password, salt, nickname, mobile, sex, status,
                    dept_id, operation_center_id, warehouse_id,
                    creator, create_time, deleted, tenant_id
                ) VALUES (
                    %s, %s, '', %s, %s, 1, 0,
                    %s, %s, %s,
                    '-1', NOW(), 0, 1
                )
            ''', (
                bd_data['bd_phone'], password_hash, bd_data['bd_name'], bd_data['bd_phone'],
                dept_id, operation_center_id, warehouse_id
            ))
            db.commit()
            return True
    except Exception as e:
        print(f"    ❌ 创建失败: {e}")
        return False
    finally:
        db.close()


def create_bd_accounts(excel_path, test_count=0):
    """批量创建BD账号"""
    print("=" * 60)
    print("Step 1: 创建BD账号")
    print("=" * 60)
    
    # 1. 读取Excel
    bd_list = read_excel(excel_path)
    print(f"\nExcel中的BD: {len(bd_list)} 个")
    
    if test_count > 0:
        bd_list = bd_list[:test_count]
        print(f"测试模式: 只处理前 {test_count} 个")
    
    # 2. 检查已存在的BD
    bd_phones = [bd['bd_phone'] for bd in bd_list]
    existing = check_existing_bd(bd_phones)
    print(f"已存在: {len(existing)} 个")
    
    # 3. 创建BD账号
    created = 0
    skipped = 0
    failed = 0
    
    for i, bd in enumerate(bd_list):
        if bd['bd_phone'] in existing:
            skipped += 1
            continue
        
        result = create_bd_account(bd)
        if result:
            created += 1
            if (created) % 10 == 0:
                print(f"  进度: {i+1}/{len(bd_list)}, 创建: {created}")
        else:
            failed += 1
        
        time.sleep(0.1)
    
    # 4. 输出汇总
    print(f"\n{'='*60}")
    print("BD账号创建汇总")
    print(f"{'='*60}")
    print(f"总计: {len(bd_list)}")
    print(f"已存在: {skipped}")
    print(f"新创建: {created}")
    print(f"失败: {failed}")
    
    return {"total": len(bd_list), "existing": skipped, "created": created, "failed": failed}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="创建BD账号")
    parser.add_argument("--excel", required=True, help="Excel文件路径")
    parser.add_argument("--test", type=int, default=0, help="测试模式：只处理前N个")
    args = parser.parse_args()
    
    if not os.path.exists(args.excel):
        print(f"❌ 文件不存在: {args.excel}")
        sys.exit(1)
    
    create_bd_accounts(args.excel, args.test)
