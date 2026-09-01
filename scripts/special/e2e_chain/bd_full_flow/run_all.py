"""
BD签约全流程 - 主入口（BD维度）

流程:
1. 从BD Excel获取下一个BD的数据
2. 校验BD账户状态（注册？分拣中心正确？）
3. 查询该BD下需要签约的网点（从网点Excel匹配）
4. 如果有需要签约的网点:
   a. 创建线索
   b. 签约线索
5. 处理完毕，继续下一个BD
6. 如果没有需要签约的站点，跳过该BD

使用方式:
  python run_all.py \
    --excel "/path/to/网点数据.xlsx" \
    --bd-excel "/path/to/BD数据.xlsx"

  # 测试模式（只处理前5个BD）
  python run_all.py \
    --excel "/path/to/网点数据.xlsx" \
    --bd-excel "/path/to/BD数据.xlsx" \
    --test 5
"""
import sys
import os
import json
import argparse
import time
from collections import defaultdict

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, PROJECT_ROOT)

import pymysql
import requests
from openpyxl import load_workbook
from config import CENTER_MAP, PROVINCE_CODE_MAP, COMMON_HEADERS, SUPER_ADMIN, API_BASE_URL as APP_URL
from geocoder import geocode_address


# ============================================================
# 工具函数
# ============================================================

def get_db():
    """获取数据库连接"""
    return pymysql.connect(
        host='rm-bp1kmprsfdog024fsro.mysql.rds.aliyuncs.com',
        port=3306, user='sf_fht_dev', password='8HUvyZf6X&FNR%5',
        database='fht_yhs', charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor, connect_timeout=5
    )


def load_bd_excel(bd_excel_path):
    """加载BD Excel数据（解析层级结构）"""
    wb = load_workbook(bd_excel_path, read_only=True)
    ws = wb.active
    
    bd_list = []
    current_center = ""
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[1] and str(row[1]).strip():
            current_center = str(row[1]).strip()
        
        phone = str(row[6]).strip() if row[6] else ""
        name = str(row[4]).strip() if row[4] else ""
        position = str(row[5]).strip() if row[5] else ""
        warehouse = str(row[3]).strip() if row[3] else ""
        
        if phone and len(phone) == 11 and phone.startswith('1'):
            bd_list.append({
                "phone": phone,
                "name": name,
                "position": position,
                "center": current_center,
                "warehouse": warehouse,
            })
    
    wb.close()
    return bd_list


def load_station_excel(excel_path):
    """加载网点Excel数据"""
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    # 按BD手机号分组
    station_by_bd = defaultdict(list)
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or not row[20] or not row[22]:
            continue
        
        phone = str(row[22]).strip() if row[22] else ""
        if not (len(phone) == 11 and phone.startswith('1')):
            continue
        
        bd_phone = str(row[29]).strip() if row[29] else ""
        if not bd_phone:
            continue
        
        # 解析省份编码
        province = str(row[23]).strip() if row[23] else ""
        province_code = PROVINCE_CODE_MAP.get(province, "")
        if not province_code:
            for key, code in PROVINCE_CODE_MAP.items():
                if key in province or province in key:
                    province_code = code
                    break
        
        station_by_bd[bd_phone].append({
            "row": row_idx,
            "station_name": str(row[20]).strip() if row[20] else "",
            "contact_name": str(row[21]).strip() if row[21] else "",
            "contact_phone": phone,
            "province": province,
            "province_code": province_code or "330000",
            "city": str(row[24]).strip() if row[24] else "",
            "district": str(row[25]).strip() if row[25] else "",
            "street": str(row[26]).strip() if row[26] else "",  # 列27: 所属街道/村/社区
            "detail_address": str(row[27]).strip() if row[27] else "",
            "station_type": 1 if str(row[4]).strip() == "门店" else 2 if str(row[4]).strip() == "回收机" else 3,
            "industry": str(row[5]).strip() if row[5] else "默认",
            "recycle_price": float(row[6]) if row[6] else 0.5,
            "clear_price": float(row[13]) if row[13] else 1.0,
            "clear_income": float(row[14]) if row[14] else 0,  # 列15: 清运收益
            "withdraw_type": str(row[16]).strip() if row[16] else "",  # 列17: 提现方式
            "bd_name": str(row[28]).strip() if row[28] else "",
            "bd_phone": bd_phone,
        })
    
    wb.close()
    return station_by_bd


def check_bd_account(phone):
    """校验BD账户状态"""
    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            SELECT id, nickname, operation_center_id, deleted 
            FROM system_users 
            WHERE mobile = %s
        """, (phone,))
        row = cur.fetchone()
    db.close()
    
    if not row:
        return {"exists": False, "msg": "BD账号不存在"}
    
    if row['deleted'] == b'\x01':
        return {"exists": False, "msg": "BD账号已删除"}
    
    if not row['operation_center_id']:
        return {"exists": True, "valid": False, "msg": "BD未绑定分拣中心"}
    
    return {
        "exists": True,
        "valid": True,
        "user_id": row['id'],
        "nickname": row['nickname'],
        "operation_center_id": row['operation_center_id'],
    }


def login_bd(phone):
    """BD登录"""
    try:
        resp = requests.post(
            f"{APP_URL}/admin-api/system/auth/sms-login",
            json={"mobile": phone, "code": "9999"},
            headers=COMMON_HEADERS,
            timeout=10,
        )
        data = resp.json()
        if data.get("code") == 0:
            return data["data"]["accessToken"]
        return None
    except:
        return None


def get_super_admin_token():
    """获取超级管理员token（使用SMS登录）"""
    resp = requests.post(
        f"{APP_URL}/admin-api/system/auth/sms-login",
        json={"mobile": "13368130837", "code": "9999"},
        headers=COMMON_HEADERS,
        timeout=10,
    )
    data = resp.json()
    if data.get("code") == 0:
        return data["data"]["accessToken"]
    return None


def create_clue(token, station_data, center_id):
    """创建线索"""
    headers = COMMON_HEADERS.copy()
    headers["Authorization"] = f"Bearer {token}"
    
    detail_address = station_data['detail_address']
    if not detail_address:
        detail_address = f"{station_data['province']}{station_data['city']}{station_data['district']}"
    
    # 地理编码
    geo_result = geocode_address(
        detail_address,
        station_data.get('province', ''),
        station_data.get('city', ''),
        station_data.get('district', ''),
    )
    
    body = {
        "poolType": 0,
        "clueName": station_data['station_name'],
        "contactName": station_data['contact_name'],
        "contactPhone": station_data['contact_phone'],
        "stationType": station_data['station_type'],
        "detailAddress": detail_address,
        "province": station_data['province'] or "浙江省",
        "provinceCode": geo_result.get("province_code") or station_data.get('province_code', "330000"),
        "city": station_data['city'] or "杭州市",
        "cityCode": geo_result.get("city_code") or "331000",
        "district": station_data['district'] or "滨江区",
        "districtCode": geo_result.get("district_code") or "330108",
        "lat": geo_result.get("lat", 0),
        "lon": geo_result.get("lon", 0),
        "belongCenterId": center_id,
    }
    
    try:
        resp = requests.post(
            f"{APP_URL}/admin-api/recycle/station-clue/create",
            json=body,
            headers=headers,
            timeout=10,
        )
        return resp.json()
    except Exception as e:
        return {"code": -1, "msg": str(e)}


def claim_clue(token, clue_id):
    """领取线索"""
    headers = COMMON_HEADERS.copy()
    headers["Authorization"] = f"Bearer {token}"
    try:
        resp = requests.post(
            f"{APP_URL}/admin-api/recycle/station-clue/claim?id={clue_id}",
            headers=headers,
            timeout=10,
        )
        return resp.json()
    except Exception as e:
        return {"code": -1, "msg": str(e)}


def sign_station(token, clue_id, station_data):
    """签约站点"""
    headers = COMMON_HEADERS.copy()
    headers["Authorization"] = f"Bearer {token}"
    
    # 地理编码
    geo_result = geocode_address(
        station_data.get("detail_address", ""),
        station_data.get("province", ""),
        station_data.get("city", ""),
        station_data.get("district", ""),
    )
    
    # 行业映射
    from config import INDUSTRY_MAP
    industry_name = station_data.get("industry", "默认")
    industry_info = INDUSTRY_MAP.get(industry_name, INDUSTRY_MAP["默认"])
    
    body = {
        "clueId": clue_id,
        "type": station_data.get("station_type", 31),
        "industry": industry_info["value"],
        "industryKey": industry_info["key"],
        "operationMode": "自营",
        "recyclePrice": station_data.get("recycle_price", 0.5),
        "clearPrice": station_data.get("clear_price", 1.0),
        "stationName": station_data["station_name"],
        "contactName": station_data["contact_name"],
        "contactPhone": station_data["contact_phone"],
        "province": station_data["province"],
        "provinceCode": geo_result.get("province_code") or station_data.get("province_code", "330000"),
        "city": station_data["city"],
        "district": station_data["district"],
        "street": station_data.get("street", ""),
        "detailAddress": station_data["detail_address"],
        "lat": geo_result.get("lat", 0),
        "lon": geo_result.get("lon", 0),
        "paymentType": 10,
        "cleartMode": 10,
        "callMode": 10,
        "minWeight": 50,
        "settlementType": 10,
        "incomeType": 1,
        "invoiceType": 1,
        "withdrawType": 1,
    }
    
    try:
        resp = requests.post(
            f"{APP_URL}/admin-api/recycle/station-clue/sign-submit",
            json=body,
            headers=headers,
            timeout=30,
        )
        return resp.json()
    except Exception as e:
        return {"code": -1, "msg": str(e)}


def find_clue(phone):
    """查询线索"""
    db = get_db()
    with db.cursor() as cur:
        cur.execute('''
            SELECT id, status, pool_type, receive_user_id, belong_center_id
            FROM station_clue 
            WHERE contact_phone = %s AND deleted = 0
        ''', (phone,))
        row = cur.fetchone()
    db.close()
    return row


# ============================================================
# 主流程
# ============================================================

def run_flow(excel_path, bd_excel_path, test_count=0):
    """BD维度的全流程"""
    print("=" * 70)
    print("BD签约全流程（BD维度）")
    print("=" * 70)
    
    # 1. 加载数据
    bd_list = load_bd_excel(bd_excel_path)
    station_by_bd = load_station_excel(excel_path)
    
    print(f"\nBD总数: {len(bd_list)}")
    print(f"有网点数据的BD: {len(station_by_bd)}")
    
    # 2. 统计可处理的BD
    processable_bds = []
    for bd in bd_list:
        if bd['phone'] in station_by_bd:
            stations = station_by_bd[bd['phone']]
            processable_bds.append({
                "bd": bd,
                "stations": stations,
            })
    
    print(f"可处理的BD: {len(processable_bds)}")
    
    if test_count > 0:
        processable_bds = processable_bds[:test_count]
        print(f"测试模式: 只处理前 {test_count} 个BD")
    
    # 3. 获取超级管理员token
    admin_token = get_super_admin_token()
    if not admin_token:
        print("❌ 获取超级管理员token失败")
        return
    print("✅ 超级管理员登录成功")
    
    # 4. 逐个处理BD
    total_success = 0
    total_fail = 0
    total_skip = 0
    
    for bd_idx, bd_item in enumerate(processable_bds):
        bd = bd_item['bd']
        stations = bd_item['stations']
        
        print(f"\n{'='*70}")
        print(f"BD [{bd_idx+1}/{len(processable_bds)}] {bd['name']} ({bd['phone']})")
        print(f"  所属分拣中心: {bd['center']}")
        print(f"  需要签约的网点: {len(stations)} 个")
        print(f"{'='*70}")
        
        # 4.1 校验BD账户
        bd_info = check_bd_account(bd['phone'])
        if not bd_info['exists']:
            print(f"  ❌ BD账号不存在，跳过")
            total_fail += 1
            continue
        
        if not bd_info.get('valid', True):
            print(f"  ❌ BD账号无效: {bd_info['msg']}，跳过")
            total_fail += 1
            continue
        
        print(f"  ✅ BD账号校验通过 (user_id={bd_info['user_id']})")
        
        # 4.2 登录BD
        token = login_bd(bd['phone'])
        if not token:
            print(f"  ❌ BD登录失败，跳过")
            total_fail += 1
            continue
        print(f"  ✅ BD登录成功")
        
        # 4.3 处理每个网点
        bd_success = 0
        bd_fail = 0
        bd_skip = 0
        
        for station_idx, station in enumerate(stations):
            print(f"\n    [{station_idx+1}/{len(stations)}] {station['station_name']} ({station['contact_phone']})")
            
            # 查询线索
            clue = find_clue(station['contact_phone'])
            if clue:
                if clue['status'] == 41:
                    print(f"      ⏭️ 已签约，跳过")
                    bd_skip += 1
                    total_skip += 1
                    continue
                if clue['status'] not in (10, 20):
                    print(f"      ⏭️ 状态={clue['status']}，跳过")
                    bd_skip += 1
                    total_skip += 1
                    continue
            
            # 创建线索（如果不存在）
            if not clue:
                print(f"      创建线索...")
                create_result = create_clue(token, station, bd_info['operation_center_id'])
                if create_result.get('code') != 0:
                    print(f"      ❌ 创建线索失败: {create_result.get('msg')}")
                    bd_fail += 1
                    total_fail += 1
                    continue
                
                clue_id = create_result.get('data')
                print(f"      ✅ 线索创建成功 (id={clue_id})")
                
                # 新创建的线索需要 claim
                print(f"      领取线索...")
                claim_result = claim_clue(token, clue_id)
                if claim_result.get('code') != 0:
                    print(f"      ❌ 领取失败: {claim_result.get('msg')}")
                    bd_fail += 1
                    total_fail += 1
                    continue
                print(f"      ✅ 领取成功")
            else:
                clue_id = clue['id']
                # 如果是status=10，需要claim
                if clue['status'] == 10:
                    print(f"      领取线索...")
                    claim_result = claim_clue(token, clue_id)
                    if claim_result.get('code') != 0:
                        print(f"      ❌ 领取失败: {claim_result.get('msg')}")
                        bd_fail += 1
                        total_fail += 1
                        continue
                    print(f"      ✅ 领取成功")
            
            # 签约
            print(f"      签约中...")
            sign_result = sign_station(token, clue_id, station)
            if sign_result.get('code') == 0:
                sign_data = sign_result.get('data', {})
                print(f"      ✅ 签约成功 (stationId={sign_data.get('stationId')})")
                bd_success += 1
                total_success += 1
            else:
                print(f"      ❌ 签约失败: {sign_result.get('msg')}")
                bd_fail += 1
                total_fail += 1
            
            time.sleep(0.5)
        
        print(f"\n  BD {bd['name']} 完成: 成功={bd_success}, 失败={bd_fail}, 跳过={bd_skip}")
    
    # 5. 汇总
    print("\n" + "=" * 70)
    print("全流程执行完成")
    print("=" * 70)
    print(f"可处理BD: {len(processable_bds)}")
    print(f"成功签约: {total_success}")
    print(f"失败: {total_fail}")
    print(f"跳过: {total_skip}")


def main():
    parser = argparse.ArgumentParser(description="BD签约全流程（BD维度）")
    parser.add_argument("--excel", required=True, help="网点Excel文件路径")
    parser.add_argument("--bd-excel", required=True, help="BD Excel文件路径")
    parser.add_argument("--test", type=int, default=0, help="测试模式：只处理前N个BD")
    args = parser.parse_args()
    
    if not os.path.exists(args.excel):
        print(f"❌ 网点文件不存在: {args.excel}")
        sys.exit(1)
    
    if not os.path.exists(args.bd_excel):
        print(f"❌ BD文件不存在: {args.bd_excel}")
        sys.exit(1)
    
    run_flow(args.excel, args.bd_excel, args.test)


if __name__ == "__main__":
    main()
