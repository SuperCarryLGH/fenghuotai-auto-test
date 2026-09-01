"""
BD签约批量导入脚本（按Excel逐条处理）

使用方式:
  # 测试前10条
  python batch_bd_sign.py --test 10

  # 全量执行
  python batch_bd_sign.py

  # 从指定索引开始
  python batch_bd_sign.py --start 100
"""
import sys
import os
import json
import time
import argparse
import requests
from collections import defaultdict

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, PROJECT_ROOT)

from openpyxl import load_workbook
import pymysql

# ============================================================
# 配置
# ============================================================
COMMON_HEADERS = {
    "Content-Type": "application/json",
    "tenant-id": "1",
    "appId": "admin",
    "sign": "admin",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
}

INDUSTRY_MAP = {
    "以车代库-自营": {"key": 11, "value": "以车代库-自营"},
    "以车代库-非自营": {"key": 10, "value": "以车代库-非自营"},
    "物业代收": {"key": 9, "value": "物业代收"},
    "其他-非垂直门店": {"key": 8, "value": "其他-非垂直门店"},
    "基层政府组织": {"key": 7, "value": "基层政府组织"},
    "服装店": {"key": 6, "value": "服装店"},
    "夫妻店": {"key": 5, "value": "夫妻店"},
    "废品回收站": {"key": 4, "value": "废品回收站"},
    "干洗店": {"key": 3, "value": "干洗店"},
    "快递驿站": {"key": 2, "value": "快递驿站"},
    "默认": {"key": 1, "value": "默认"},
}

PROVINCE_CODE_MAP = {
    "浙江省": "330000", "江苏省": "320000", "安徽省": "340000",
    "湖北省": "420000", "湖南省": "430000", "四川省": "510000",
    "重庆市": "500000", "广西": "450000", "广西壮族自治区": "450000",
    "广东省": "440000", "河南省": "410000", "河北省": "130000",
}

RESULT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "batch_sign_result.json")

# ============================================================
# 工具函数
# ============================================================
def get_db():
    return pymysql.connect(
        host='rm-bp1kmprsfdog024fsro.mysql.rds.aliyuncs.com',
        port=3306, user='sf_fht_dev', password='8HUvyZf6X&FNR%5',
        database='fht_yhs', charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor, connect_timeout=5
    )

def read_excel(excel_path):
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=2):
        if not row[0] or not row[20]:
            continue
        industry_name = str(row[4]) if row[4] else "默认"
        industry_info = INDUSTRY_MAP.get(industry_name, INDUSTRY_MAP["默认"])
        province_name = str(row[23]) if row[23] else ""
        province_code = PROVINCE_CODE_MAP.get(province_name, "")
        if not province_code:
            for key, code in PROVINCE_CODE_MAP.items():
                if key in province_name or province_name in key:
                    province_code = code
                    break
        data.append({
            "index": row_idx,
            "stationName": str(row[20]) if row[20] else "",
            "contactName": str(row[21]) if row[21] else "",
            "contactPhone": str(row[22]) if row[22] else "",
            "bdPhone": str(row[29]) if row[29] else "",
            "industry": industry_info["value"],
            "industryKey": industry_info["key"],
            "operationMode": str(row[5]) if row[5] else "自营",
            "recyclePrice": float(row[6]) if row[6] else 0,
            "clearPrice": float(row[13]) if row[13] else 0,
            "province": province_name,
            "provinceCode": province_code,
            "city": str(row[24]) if row[24] else "",
            "district": str(row[25]) if row[25] else "",
            "street": str(row[26]) if row[26] else "",
            "detailAddress": str(row[27]) if row[27] else "",
        })
    wb.close()
    return data

def login(mobile):
    try:
        resp = requests.post("https://api-fht-dev.hengyishou.com/admin-api/system/auth/sms-login",
            json={"mobile": mobile, "code": "9999"}, headers=COMMON_HEADERS, timeout=10)
        data = resp.json()
        return data["data"]["accessToken"] if data.get("code") == 0 else None
    except:
        return None

def find_clue(phone):
    db = get_db()
    with db.cursor() as cur:
        cur.execute('SELECT id, status, receive_user_id FROM station_clue WHERE contact_phone=%s AND deleted=0', (phone,))
        row = cur.fetchone()
    db.close()
    return row

def get_bd_id(phone):
    db = get_db()
    with db.cursor() as cur:
        cur.execute('SELECT id, nickname FROM system_users WHERE mobile=%s', (phone,))
        row = cur.fetchone()
    db.close()
    return (row['id'], row['nickname']) if row else (None, None)

def claim_clue(token, clue_id):
    """领取线索（status=10 → status=20）"""
    headers = COMMON_HEADERS.copy()
    headers["Authorization"] = f"Bearer {token}"
    try:
        resp = requests.post(
            f"https://api-fht-dev.hengyishou.com/admin-api/recycle/station-clue/claim?id={clue_id}",
            headers=headers, timeout=10)
        return resp.json()
    except:
        return {"code": -1}

def sign(token, clue_id, data):
    headers = COMMON_HEADERS.copy()
    headers["Authorization"] = f"Bearer {token}"
    body = {
        "clueId": clue_id, "type": 31,
        "industry": data["industry"], "industryKey": data["industryKey"],
        "operationMode": data["operationMode"],
        "recyclePrice": data["recyclePrice"], "clearPrice": data["clearPrice"],
        "stationName": data["stationName"],
        "contactName": data["contactName"], "contactPhone": data["contactPhone"],
        "province": data["province"], "provinceCode": data["provinceCode"],
        "city": data["city"], "district": data["district"],
        "street": data["street"], "detailAddress": data["detailAddress"],
        "paymentType": 10, "cleartMode": 10, "callMode": 10,
        "minWeight": 50, "settlementType": 10, "incomeType": 1, "invoiceType": 1, "withdrawType": 1,
    }
    try:
        resp = requests.post("https://api-fht-dev.hengyishou.com/admin-api/recycle/station-clue/sign-submit",
            json=body, headers=headers, timeout=30)
        return resp.json()
    except Exception as e:
        return {"code": -1, "msg": str(e)}

def load_result():
    if os.path.exists(RESULT_FILE):
        with open(RESULT_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"success": [], "failed": [], "last_index": 0}

def save_result(result):
    with open(RESULT_FILE, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

# ============================================================
# 主流程
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="BD签约批量导入")
    parser.add_argument("--excel", default="/Users/rs/Documents/20260827 逸掌柜App网点账号开通信息收集表（修改）.xlsx")
    parser.add_argument("--start", type=int, default=0)
    parser.add_argument("--test", type=int, default=0, help="测试模式：只处理前N条")
    args = parser.parse_args()

    print("=" * 60)
    print("BD签约批量导入（按Excel逐条处理）")
    print("=" * 60)

    all_data = read_excel(args.excel)
    print(f"\n读取Excel: {len(all_data)} 条")

    # 加载历史结果
    result = load_result()
    success_phones = set(item.get("phone") for item in result["success"])
    fail_phones = set(item.get("phone") for item in result["failed"])
    processed = success_phones | fail_phones
    print(f"已处理: {len(processed)} 条")

    total_success = len(result["success"])
    total_fail = len(result["failed"])
    total_skip = 0
    tokens = {}
    count = 0

    for i, record in enumerate(all_data):
        if record["contactPhone"] in processed:
            continue
        if args.start and record["index"] < args.start:
            continue
        if args.test > 0 and count >= args.test:
            break

        # 查线索
        clue = find_clue(record["contactPhone"])
        if not clue:
            total_fail += 1
            result["failed"].append({"phone": record["contactPhone"], "station": record["stationName"], "error": "未找到线索"})
            count += 1
            continue

        if clue["status"] == 41:
            total_skip += 1
            count += 1
            continue

        if clue["status"] not in (10, 20):
            total_skip += 1
            count += 1
            continue

        # 登录BD
        bd_phone = record["bdPhone"]
        if bd_phone not in tokens:
            token = login(bd_phone)
            if not token:
                total_fail += 1
                result["failed"].append({"phone": record["contactPhone"], "station": record["stationName"], "error": "登录失败"})
                count += 1
                continue
            tokens[bd_phone] = token

        # 获取BD user_id
        bd_id, bd_name = get_bd_id(bd_phone)
        if not bd_id:
            total_fail += 1
            result["failed"].append({"phone": record["contactPhone"], "station": record["stationName"], "error": "未找到BD账号"})
            count += 1
            continue

        # 如果线索是status=10（待处理），先领取
        if clue["status"] == 10:
            print(f"  [{count+1}] 领取线索...")
            claim_r = claim_clue(tokens[bd_phone], clue["id"])
            if claim_r.get("code") != 0:
                total_fail += 1
                result["failed"].append({"phone": record["contactPhone"], "station": record["stationName"], "error": f"领取失败: {claim_r.get('msg')}"})
                count += 1
                continue

        # 签约
        r = sign(tokens[bd_phone], clue["id"], record)
        if r.get("code") == 0:
            d = r.get("data", {})
            total_success += 1
            result["success"].append({
                "phone": record["contactPhone"], "station": record["stationName"],
                "clueId": clue["id"], "signId": d.get("signId"), "stationId": d.get("stationId"),
            })
            print(f"[{count+1}] ✅ {record['stationName']} ({record['contactPhone']}) -> stationId={d.get('stationId')}")
        else:
            total_fail += 1
            result["failed"].append({
                "phone": record["contactPhone"], "station": record["stationName"],
                "error": r.get("msg", "未知错误"),
            })
            print(f"[{count+1}] ❌ {record['stationName']}: {r.get('msg')}")

        count += 1
        result["last_index"] = record["index"]
        save_result(result)
        time.sleep(0.3)

    print(f"\n{'='*60}")
    print(f"汇总: 成功={total_success}, 失败={total_fail}, 跳过={total_skip}")
    print(f"结果文件: {RESULT_FILE}")
    print(f"{'='*60}")

    result["summary"] = {"success": total_success, "fail": total_fail, "skip": total_skip}
    save_result(result)

if __name__ == "__main__":
    main()
