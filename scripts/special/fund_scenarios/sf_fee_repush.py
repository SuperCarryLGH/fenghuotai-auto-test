"""
顺丰费用重推 - 从Excel批量推送(去重)到 prod
用法:
  python sf_fee_repush.py --dry-run        # 预览解析结果,不推送
  python sf_fee_repush.py --limit 1        # 只推首条(验证prod接收)
  python sf_fee_repush.py                  # 全量推送(去重,10s间隔)
"""
import sys, os, re, json, time, argparse
import requests
from openpyxl import load_workbook

EXCEL = '/Users/rs/Documents/订单费用重推.xlsx'
ENDPOINT = 'https://api-fht.hengyishou.com/app-api/recycle/express/fee-push/sf'
RESULT_FILE = '/Users/rs/Documents/顺丰费用重推结果.json'

def parse_cell(raw: str):
    """解析 B 列: sign=..., content={...} → (sign, content_json_str)"""
    raw = str(raw).strip()
    sign = None
    m = re.search(r'sign=(\S+?)[,\s]', raw)
    if m: sign = m.group(1)
    m = re.search(r'content=(\{.*\})', raw, re.DOTALL)
    if not m:
        return None, None, None
    content = m.group(1)
    try:
        cj = json.loads(content)
    except Exception:
        return None, None, None
    return sign, content, cj

def load_data():
    wb = load_workbook(EXCEL, read_only=True)
    ws = wb['Sheet1']
    items = []
    seen = set()
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if not row or not row[0]:
            continue
        sign, content, cj = parse_cell(row[1])
        if cj is None:
            print(f'  ⚠️ 第{i}行解析失败,跳过: {row[0]}')
            continue
        waybill = cj.get('waybillNo', row[0])
        if waybill in seen:
            print(f'  ⏭️ 去重跳过: {waybill} (第{i}行)')
            continue
        seen.add(waybill)
        items.append({'excel_row': i, 'waybill': waybill, 'orderNo': cj.get('orderNo', ''),
                      'sign': sign, 'content': content})
    wb.close()
    return items

def post(content, sign):
    data = {"content": content}
    if sign: data["sign"] = sign
    try:
        r = requests.post(ENDPOINT, data=data,
            headers={"Content-Type": "application/x-www-form-urlencoded", "User-Agent": "Mozilla/5.0"},
            timeout=30)
        if r.status_code == 200:
            try:
                return r.json()
            except Exception:
                return {"raw": r.text}
        return {"http": r.status_code, "raw": r.text[:200]}
    except Exception as e:
        return {"error": str(e)}

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--limit', type=int, default=0)
    ap.add_argument('--start', type=int, default=1, help='从第N条开始(1-based)')
    ap.add_argument('--interval', type=int, default=10)
    args = ap.parse_args()

    items = load_data()
    print(f'数据: 共解析 {len(items)} 条(已去重) | 目标: {ENDPOINT} | 间隔: {args.interval}s')
    if args.start > 1:
        items = items[args.start-1:]
        print(f'从第 {args.start} 条开始(跳过已推前 {args.start-1} 条)')
    if args.limit > 0:
        items = items[:args.limit]
        print(f'仅推送前 {args.limit} 条')
    if args.dry_run:
        print('\n[DRY-RUN] 预览:')
        for it in items:
            print(f'  [{it["excel_row"]}] {it["waybill"]} {it["orderNo"]} sign={str(it["sign"])[:20]}...')
        return

    results = []
    for idx, it in enumerate(items, 1):
        print(f'\n[{idx}/{len(items)}] 运单 {it["waybill"]} 订单 {it["orderNo"]}')
        resp = post(it['content'], it['sign'])
        code = resp.get('code') if isinstance(resp, dict) else None
        if code == 200:
            print(f'  ✅ code=200 接收成功')
            results.append({'waybill': it['waybill'], 'orderNo': it['orderNo'], 'row': it['excel_row'], 'result': 'success'})
        else:
            print(f'  ❌ 失败: {json.dumps(resp, ensure_ascii=False)[:200]}')
            results.append({'waybill': it['waybill'], 'orderNo': it['orderNo'], 'row': it['excel_row'], 'result': 'fail', 'resp': str(resp)[:200]})
        if idx < len(items):
            print(f'  等待 {args.interval}s...')
            time.sleep(args.interval)

    succ = sum(1 for r in results if r['result'] == 'success')
    print(f'\n{"="*55}\n完成: 总{len(results)} 成功{succ} 失败{len(results)-succ}')
    with open(RESULT_FILE, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f'结果已存: {RESULT_FILE}')

if __name__ == '__main__':
    main()