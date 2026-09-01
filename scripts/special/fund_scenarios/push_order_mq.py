"""
苏周到订单 - push-order-mq 推送(线上)
用法:
  python push_order_mq.py --dry-run   # 预览
  python push_order_mq.py             # 全量推送(10s间隔)
"""
import sys, os, time, json, argparse
import requests
from openpyxl import load_workbook

EXCEL = '/Users/rs/Documents/苏周到订单.xlsx'
BASE = 'https://api-fht.hengyishou.com/app-api/recycle/express/push-order-mq'
RESULT_FILE = '/Users/rs/Documents/苏周到推送结果.json'

HEADERS = {"User-Agent": "Mozilla/5.0", "tenant-id": "1"}

def load():
    wb = load_workbook(EXCEL, read_only=True)
    items = []
    for sheet, source in [('取消', 'cancel'), ('完成', 'complete')]:
        ws = wb[sheet]
        for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not r or not r[0]:
                continue
            items.append({'orderId': str(r[0]).strip(), 'bizSource': source,
                          'waybill': str(r[1]).strip() if r[1] else '',
                          'orderNo': str(r[4]).strip() if len(r) > 4 and r[4] else '',
                          'excel': f'{sheet}!R{i}'})
    wb.close()
    return items

def call(order_id, source):
    url = f'{BASE}?orderId={order_id}&bizSource={source}'
    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        if r.status_code == 200:
            try:
                b = r.json()
                return b.get('code'), str(b.get('data', ''))
            except Exception:
                return -1, f'非JSON: {r.text[:100]}'
        return -1, f'HTTP {r.status_code}'
    except Exception as e:
        return -1, str(e)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--interval', type=int, default=10)
    args = ap.parse_args()

    items = load()
    print(f'数据: {len(items)} 条 (取消 {sum(1 for i in items if i["bizSource"]=="cancel")} / 完成 {sum(1 for i in items if i["bizSource"]=="complete")})')
    print(f'目标: {BASE} | 间隔: {args.interval}s | tenant-id: 1')

    if args.dry_run:
        print('[DRY-RUN] 预览:')
        for it in items[:5]:
            print(f'  [{it["bizSource"]}] {it["orderId"]} {it["orderNo"]} {it["waybill"]} ({it["excel"]})')
        return

    results = []
    for idx, it in enumerate(items, 1):
        code, data = call(it['orderId'], it['bizSource'])
        if code == 0 and data == '推送成功':
            cls = 'success'
        elif code == 0 and '未找到' in data:
            cls = 'notfound'
        else:
            cls = 'fail'
        print(f'[{idx}/{len(items)}] {it["bizSource"]} {it["orderId"]} → {cls} ({data})')
        results.append({'orderId': it['orderId'], 'bizSource': it['bizSource'],
                        'orderNo': it['orderNo'], 'waybill': it['waybill'],
                        'excel': it['excel'], 'result': cls, 'data': data})
        if idx < len(items):
            time.sleep(args.interval)

    from collections import Counter
    c = Counter(r['result'] for r in results)
    print(f'\n{"="*55}\n完成: {len(results)} | 成功{c["success"]} | 未找到{c["notfound"]} | 失败{c["fail"]}')
    with open(RESULT_FILE, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f'结果已存: {RESULT_FILE}')

if __name__ == '__main__':
    main()