import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

with open("/Users/rs/PycharmProjects/PythonProject1/openapi.json") as f:
    spec = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "API文档"

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4472C4")
tag_fill = PatternFill("solid", fgColor="D9E2F3")
tag_font = Font(bold=True, size=12)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = ["模块", "接口名称", "请求方法", "接口路径", "参数说明"]
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 55
ws.column_dimensions['E'].width = 60

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

row = 2
server_url = spec.get("servers", [{}])[0].get("url", "")

paths = spec.get("paths", {})
for path, methods in paths.items():
    for method, detail in methods.items():
        tags = detail.get("tags", ["未分类"])
        summary = detail.get("summary", "")
        params_desc = []

        parameters = detail.get("parameters", [])
        for p in parameters:
            name = p.get("name", "")
            required = "必填" if p.get("required") else "选填"
            desc = p.get("description", "")
            params_desc.append(f"{name}({required}) - {desc}")

        request_body = detail.get("requestBody", {})
        if request_body:
            content = request_body.get("content", {})
            for media_type, media_detail in content.items():
                schema = media_detail.get("schema", {})
                ref = schema.get("$ref", schema.get("properties", {}))
                if isinstance(ref, str):
                    params_desc.append(f"Body: {ref.split('/')[-1]}")
                elif ref:
                    props = []
                    for k, v in ref.items():
                        desc = v.get("description", "")
                        props.append(f"{k}({v.get('type', '')}) - {desc}")
                    if props:
                        params_desc.append(f"Body: {', '.join(props)}")

        tag = tags[0] if tags else "未分类"
        ws.cell(row=row, column=1, value=tag).border = thin_border
        ws.cell(row=row, column=2, value=summary).border = thin_border
        ws.cell(row=row, column=3, value=method.upper()).border = thin_border
        ws.cell(row=row, column=4, value=f"{path}").border = thin_border
        ws.cell(row=row, column=5, value="\n".join(params_desc) if params_desc else "").border = thin_border

        for c in range(1, 6):
            ws.cell(row=row, column=c).alignment = Alignment(vertical='top', wrap_text=True)
        row += 1

output_path = "/Users/rs/PycharmProjects/PythonProject1/API文档.xlsx"
wb.save(output_path)
print(f"OK -> {output_path}, 共 {row - 2} 条接口")
